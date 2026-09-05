import os
import io
import csv
import tempfile
import uuid
import zipfile
from functools import wraps
from datetime import datetime
from mimetypes import guess_type

import psycopg2
from psycopg2.extras import RealDictCursor

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    session,
    send_file
)

from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

from supabase import create_client

from openpyxl import load_workbook
from odf.opendocument import load as load_ods_document
from odf.table import (
    Table as ODFTable,
    TableRow as ODFTableRow,
    TableCell as ODFTableCell
)
from odf.text import P as ODFParagraph


ADMIN_NAME = "Admin"
ADMIN_PASSWORD = "fola2004"

app = Flask(__name__)

app.secret_key = os.environ.get(
    "SECRET_KEY",
    "change-this-secret-key"
)

app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024


DATABASE_URL = os.environ.get("DATABASE_URL")
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY")

SUPABASE_BUCKET = "school files"


if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL is not set.")

if not SUPABASE_URL:
    raise RuntimeError("SUPABASE_URL is not set.")

if not SUPABASE_KEY:
    raise RuntimeError("SUPABASE_KEY is not set.")


supabase = create_client(
    SUPABASE_URL,
    SUPABASE_KEY
)


ALLOWED_RESULT_EXTENSIONS = {
    "pdf",
    "xlsx",
    "xlsm",
    "xls",
    "ods",
    "csv",
    "txt",
    "png",
    "jpg",
    "jpeg",
    "gif",
    "webp"
}


TERMS = [
    "First Term",
    "Second Term",
    "Third Term"
]


CLASSES = [
    "JSS1",
    "JSS2",
    "JSS3",
    "SS1",
    "SS2",
    "SS3"
]


DEPARTMENTS = [
    "Science",
    "Art",
    "Commercial"
]


def get_db():
    return psycopg2.connect(
        DATABASE_URL,
        cursor_factory=RealDictCursor
    )


def init_db():
    conn = get_db()

    try:
        cur = conn.cursor()

        cur.execute("""
            CREATE TABLE IF NOT EXISTS students (
                id SERIAL PRIMARY KEY,
                name TEXT UNIQUE NOT NULL,
                class_name TEXT NOT NULL,
                department TEXT NOT NULL,
                password TEXT NOT NULL
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS teachers (
                id SERIAL PRIMARY KEY,
                name TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS subjects (
                id SERIAL PRIMARY KEY,
                subject_name TEXT NOT NULL,
                subject_link TEXT NOT NULL,
                class_name TEXT NOT NULL,
                department TEXT NOT NULL,
                term TEXT NOT NULL
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS results (
                id SERIAL PRIMARY KEY,
                student_id INTEGER NOT NULL,
                student_name TEXT NOT NULL,
                filename TEXT NOT NULL,
                original_filename TEXT NOT NULL,
                term TEXT NOT NULL,
                FOREIGN KEY (student_id)
                REFERENCES students(id)
                ON DELETE CASCADE
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS news (
                id SERIAL PRIMARY KEY,
                title TEXT NOT NULL,
                content TEXT NOT NULL,
                date TEXT NOT NULL,
                admin TEXT NOT NULL
            )
        """)

        cur.execute("""
            ALTER TABLE results
            ADD COLUMN IF NOT EXISTS term TEXT
        """)

        cur.execute("""
            ALTER TABLE results
            ADD COLUMN IF NOT EXISTS original_filename TEXT
        """)

        cur.execute("""
            ALTER TABLE results
            ADD COLUMN IF NOT EXISTS filename TEXT
        """)

        conn.commit()

        cur.close()

    finally:
        conn.close()


init_db()


def normalize_name(name):
    return " ".join(
        str(name or "").strip().split()
    )


def is_senior_class(class_name):
    return str(class_name or "").upper().startswith("SS")


def is_junior_class(class_name):
    return str(class_name or "").upper().startswith("JSS")


def allowed_result_file(filename):
    if not filename:
        return False

    if "." not in filename:
        return False

    extension = filename.rsplit(".", 1)[1].lower()

    return extension in ALLOWED_RESULT_EXTENSIONS


def storage_path(filename):
    return f"results/{filename}"


def delete_from_storage(path):
    try:
        supabase.storage.from_(
            SUPABASE_BUCKET
        ).remove([path])
    except Exception as e:
        print(
            "Storage delete error:",
            repr(e)
        )


def upload_to_storage(path, data, content_type):
    return supabase.storage.from_(
        SUPABASE_BUCKET
    ).upload(
        path=path,
        file=data,
        file_options={
            "content-type": content_type,
            "cache-control": "3600",
            "upsert": "true"
        }
    )


def download_from_storage(path):
    return supabase.storage.from_(
        SUPABASE_BUCKET
    ).download(path)


def clear_login_sessions():
    session.pop("admin", None)
    session.pop("student_id", None)
    session.pop("student_name", None)
    session.pop("class_name", None)
    session.pop("department", None)
    session.pop("teacher_id", None)
    session.pop("teacher_name", None)


def admin_required(view):
    @wraps(view)
    def wrapped_view(*args, **kwargs):
        if not session.get("admin"):
            flash("Administrator login required.")
            return redirect(url_for("login"))

        return view(*args, **kwargs)

    return wrapped_view


def student_required(view):
    @wraps(view)
    def wrapped_view(*args, **kwargs):
        if not session.get("student_id"):
            flash("Student login required.")
            return redirect(url_for("login"))

        return view(*args, **kwargs)

    return wrapped_view


def teacher_required(view):
    @wraps(view)
    def wrapped_view(*args, **kwargs):
        if not session.get("teacher_id"):
            flash("Teacher login required.")
            return redirect(url_for("login"))

        return view(*args, **kwargs)

    return wrapped_view


def find_admin(name, password):
    entered_name = normalize_name(name)

    if (
        entered_name.lower() == ADMIN_NAME.lower()
        and password == ADMIN_PASSWORD
    ):
        return ADMIN_NAME

    return None


def password_matches(stored_password, entered_password):
    if not stored_password:
        return False

    try:
        if check_password_hash(
            stored_password,
            entered_password
        ):
            return True
    except Exception:
        pass

    return stored_password == entered_password


def read_xlsx(file_bytes):
    if not file_bytes:
        raise ValueError(
            "Empty Excel file."
        )

    workbook = load_workbook(
        io.BytesIO(file_bytes),
        data_only=True
    )

    sheets = {}

    for sheet in workbook.worksheets:
        rows = []

        for row in sheet.iter_rows(
            values_only=True
        ):
            rows.append(
                list(row)
            )

        sheets[sheet.title] = rows

    if not sheets:
        raise ValueError(
            "No worksheets found."
        )

    return sheets


def _ods_cell_text(cell):
    paragraphs = []

    for paragraph in cell.getElementsByType(
        ODFParagraph
    ):
        text = ""

        for node in paragraph.childNodes:
            if hasattr(node, "data"):
                text += str(node.data)

        if text:
            paragraphs.append(text)

    value = "\n".join(paragraphs)

    repeat = cell.getAttribute(
        "numbercolumnsrepeated"
    )

    span = cell.getAttribute(
        "numbercolumnsspanned"
    )

    try:
        repeat = int(repeat or 1)
    except Exception:
        repeat = 1

    try:
        span = int(span or 1)
    except Exception:
        span = 1

    return value, repeat, span


def read_ods(file_bytes):
    if not file_bytes:
        raise ValueError(
            "Empty ODS file."
        )

    temp_path = os.path.join(
        os.getenv("TMPDIR")
        or tempfile.gettempdir(),
        f"fola_result_{uuid.uuid4().hex}.ods"
    )

    document = None

    try:
        with open(
            temp_path,
            "wb"
        ) as output:
            output.write(file_bytes)

        document = load_ods_document(
            temp_path
        )

        sheets = {}

        for table in document.spreadsheet.getElementsByType(
            ODFTable
        ):
            sheet_name = str(
                table.getAttribute("name")
                or "Sheet"
            )

            rows = []

            for row in table.childNodes:
                if (
                    not hasattr(row, "qname")
                    or row.qname != ODFTableRow.qname
                ):
                    continue

                values = []

                for cell in row.childNodes:
                    if (
                        not hasattr(cell, "qname")
                        or cell.qname != ODFTableCell.qname
                    ):
                        continue

                    value, repeat, span = _ods_cell_text(
                        cell
                    )

                    values.extend(
                        [value] *
                        max(
                            1,
                            repeat * span
                        )
                    )

                rows.append(values)

            cleaned_rows = []
            max_columns = 0

            for row in rows:
                cleaned = list(row)

                while (
                    cleaned
                    and str(cleaned[-1]).strip() == ""
                ):
                    cleaned.pop()

                if cleaned:
                    cleaned_rows.append(
                        cleaned
                    )

                    max_columns = max(
                        max_columns,
                        len(cleaned)
                    )

            for row in cleaned_rows:
                row.extend(
                    [""] *
                    (
                        max_columns
                        - len(row)
                    )
                )

            sheets[sheet_name] = cleaned_rows

        if not sheets:
            raise ValueError(
                "No sheets found in ODS file."
            )

        return sheets

    finally:
        document = None

        try:
            os.remove(temp_path)
        except OSError:
            pass


def read_csv_file(file_bytes):
    text = file_bytes.decode(
        "utf-8-sig",
        errors="replace"
    )

    reader = csv.reader(
        io.StringIO(text)
    )

    rows = []

    for row in reader:
        rows.append(
            list(row)
        )

    return {
        "CSV": rows
    }


def read_text_file(file_bytes):
    text = file_bytes.decode(
        "utf-8-sig",
        errors="replace"
    )

    rows = [
        [line]
        for line in text.splitlines()
    ]

    return {
        "Text": rows
    }


def validate_ods(file_bytes):
    if not file_bytes:
        raise ValueError(
            "Empty ODS file."
        )

    with zipfile.ZipFile(
        io.BytesIO(file_bytes)
    ) as archive:

        if "mimetype" not in archive.namelist():
            raise ValueError(
                "Invalid ODS file."
            )

        mimetype = archive.read(
            "mimetype"
        ).decode(
            "utf-8",
            errors="ignore"
        ).strip()

        if mimetype != (
            "application/vnd.oasis.opendocument.spreadsheet"
        ):
            raise ValueError(
                "The uploaded file is not a valid ODS spreadsheet."
            )


@app.route(
    "/",
    methods=["GET", "POST"]
)
@app.route(
    "/login",
    methods=["GET", "POST"]
)
def login():

    if request.method == "POST":

        user_type = request.form.get(
            "user_type",
            ""
        ).strip().lower()

        name = normalize_name(
            request.form.get(
                "name",
                ""
            )
        )

        password = request.form.get(
            "password",
            ""
        )

        if not name or not password:
            flash(
                "Please enter your name and password."
            )

            return render_template(
                "login.html"
            )

        if user_type == "admin":

            admin_name = find_admin(
                name,
                password
            )

            if admin_name:
                clear_login_sessions()

                session["admin"] = admin_name

                return redirect(
                    url_for(
                        "admin_dashboard"
                    )
                )

            flash(
                "Invalid administrator name or password."
            )

            return render_template(
                "login.html"
            )

        if user_type == "student":

            conn = get_db()

            try:
                cur = conn.cursor()

                cur.execute(
                    """
                    SELECT *
                    FROM students
                    WHERE LOWER(TRIM(name))
                          = LOWER(TRIM(%s))
                    LIMIT 1
                    """,
                    (name,)
                )

                student = cur.fetchone()

            finally:
                conn.close()

            if (
                student
                and password_matches(
                    student["password"],
                    password
                )
            ):

                clear_login_sessions()

                session["student_id"] = student["id"]
                session["student_name"] = student["name"]
                session["class_name"] = student["class_name"]
                session["department"] = student["department"]

                return redirect(
                    url_for("dashboard")
                )

            flash(
                "Invalid student name or password."
            )

            return render_template(
                "login.html"
            )

        if user_type == "teacher":

            conn = get_db()

            try:
                cur = conn.cursor()

                cur.execute(
                    """
                    SELECT *
                    FROM teachers
                    WHERE LOWER(TRIM(name))
                          = LOWER(TRIM(%s))
                    LIMIT 1
                    """,
                    (name,)
                )

                teacher = cur.fetchone()

            finally:
                conn.close()

            if (
                teacher
                and password_matches(
                    teacher["password"],
                    password
                )
            ):

                clear_login_sessions()

                session["teacher_id"] = teacher["id"]
                session["teacher_name"] = teacher["name"]

                return redirect(
                    url_for("teacher")
                )

            flash(
                "Invalid teacher name or password."
            )

            return render_template(
                "login.html"
            )

        flash(
            "Please select Student, Teacher or Administrator."
        )

    return render_template(
        "login.html"
    )


@app.route("/logout")
def logout():
    session.clear()

    return redirect(
        url_for("login")
    )


@app.route("/dashboard")
@student_required
def dashboard():

    return render_template(
        "dashboard.html",
        student_name=session.get(
            "student_name"
        ),
        class_name=session.get(
            "class_name"
        ),
        department=session.get(
            "department"
        )
    )


@app.route("/subjects")
@student_required
def subjects():

    class_name = session.get(
        "class_name"
    )

    department = session.get(
        "department"
    )

    term = request.args.get(
        "term",
        "First Term"
    ).strip()

    if not class_name:
        flash(
            "Your class information was not found. Please log in again."
        )

        session.clear()

        return redirect(
            url_for("login")
        )

    if is_junior_class(
        class_name
    ):
        department = "General"

    if term not in TERMS:
        term = "First Term"

    conn = get_db()

    try:
        cur = conn.cursor()

        cur.execute(
            """
            SELECT *
            FROM subjects
            WHERE class_name = %s
              AND department = %s
              AND term = %s
            ORDER BY subject_name
            """,
            (
                class_name,
                department,
                term
            )
        )

        subject_rows = cur.fetchall()

    finally:
        conn.close()

    return render_template(
        "subjects.html",
        subjects=subject_rows,
        class_name=class_name,
        department=department,
        term=term,
        terms=TERMS
    )


@app.route("/results")
@student_required
def results():

    student_id = session.get(
        "student_id"
    )

    term = request.args.get(
        "term",
        "All"
    ).strip()

    conn = get_db()

    try:
        cur = conn.cursor()

        if term in TERMS:

            cur.execute(
                """
                SELECT
                    id,
                    term,
                    original_filename,
                    filename
                FROM results
                WHERE student_id = %s
                  AND term = %s
                ORDER BY id DESC
                """,
                (
                    student_id,
                    term
                )
            )

        else:

            term = "All"

            cur.execute(
                """
                SELECT
                    id,
                    term,
                    original_filename,
                    filename
                FROM results
                WHERE student_id = %s
                ORDER BY id DESC
                """,
                (student_id,)
            )

        result_rows = cur.fetchall()

    finally:
        conn.close()

    return render_template(
        "results.html",
        results=result_rows,
        term=term,
        terms=TERMS
    )


@app.route(
    "/view-result/<int:result_id>"
)
@student_required
def view_result(result_id):

    student_id = session.get(
        "student_id"
    )

    conn = get_db()

    try:
        cur = conn.cursor()

        cur.execute(
            """
            SELECT *
            FROM results
            WHERE id = %s
              AND student_id = %s
            LIMIT 1
            """,
            (
                result_id,
                student_id
            )
        )

        result = cur.fetchone()

    finally:
        conn.close()

    if not result:
        flash(
            "Result not found."
        )

        return redirect(
            url_for("results")
        )

    filename = result["filename"]

    extension = os.path.splitext(
        filename
    )[1].lower()

    file_type = "file"
    sheets = None

    if extension in (
        ".xlsx",
        ".xlsm"
    ):

        try:
            file_bytes = download_from_storage(
                storage_path(filename)
            )

            sheets = read_xlsx(
                file_bytes
            )

            file_type = "spreadsheet"

        except Exception as e:

            print(
                "RESULT OPEN ERROR:",
                type(e).__name__,
                repr(e)
            )

            flash(
                "The Excel result could not be opened."
            )

            return redirect(
                url_for("results")
            )

    elif extension == ".ods":

        try:
            file_bytes = download_from_storage(
                storage_path(filename)
            )

            sheets = read_ods(
                file_bytes
            )

            file_type = "spreadsheet"

        except Exception as e:

            print(
                "RESULT OPEN ERROR:",
                type(e).__name__,
                repr(e)
            )

            flash(
                "The ODS result could not be opened."
            )

            return redirect(
                url_for("results")
            )

    elif extension == ".csv":

        try:
            file_bytes = download_from_storage(
                storage_path(filename)
            )

            sheets = read_csv_file(
                file_bytes
            )

            file_type = "spreadsheet"

        except Exception as e:

            print(
                "RESULT OPEN ERROR:",
                type(e).__name__,
                repr(e)
            )

            flash(
                "The CSV result could not be opened."
            )

            return redirect(
                url_for("results")
            )

    elif extension == ".txt":

        try:
            file_bytes = download_from_storage(
                storage_path(filename)
            )

            sheets = read_text_file(
                file_bytes
            )

            file_type = "spreadsheet"

        except Exception as e:

            print(
                "RESULT OPEN ERROR:",
                type(e).__name__,
                repr(e)
            )

            flash(
                "The text result could not be opened."
            )

            return redirect(
                url_for("results")
            )

    return render_template(
        "view_result.html",
        result=result,
        file_type=file_type,
        sheets=sheets,
        result_file=url_for(
            "result_file",
            filename=filename
        )
    )


@app.route(
    "/result-file/<path:filename>"
)
@student_required
def result_file(filename):

    student_id = session.get(
        "student_id"
    )

    conn = get_db()

    try:
        cur = conn.cursor()

        cur.execute(
            """
            SELECT *
            FROM results
            WHERE filename = %s
              AND student_id = %s
            LIMIT 1
            """,
            (
                filename,
                student_id
            )
        )

        result = cur.fetchone()

    finally:
        conn.close()

    if not result:
        return (
            "Result not found.",
            404
        )

    try:

        data = download_from_storage(
            storage_path(filename)
        )

        content_type = (
            guess_type(
                result["original_filename"]
            )[0]
            or "application/octet-stream"
        )

        return send_file(
            io.BytesIO(data),
            mimetype=content_type,
            download_name=result[
                "original_filename"
            ],
            as_attachment=False
        )

    except Exception as e:

        print(
            "RESULT FILE ERROR:",
            repr(e)
        )

        return (
            "Could not open result file.",
            500
        )


@app.route("/news")
@student_required
def news():

    conn = get_db()

    try:
        cur = conn.cursor()

        cur.execute(
            """
            SELECT *
            FROM news
            ORDER BY id DESC
            """
        )

        news_list = cur.fetchall()

    finally:
        conn.close()

    return render_template(
        "news.html",
        news=news_list,
        back_url=url_for("dashboard"),
        back_text="Dashboard"
    )


@app.route("/teacher")
@teacher_required
def teacher():

    class_name = request.args.get(
        "class_name",
        ""
    ).strip()

    department = request.args.get(
        "department",
        ""
    ).strip()

    term = request.args.get(
        "term",
        "First Term"
    ).strip()

    if term not in TERMS:
        term = "First Term"

    if class_name not in CLASSES:
        class_name = CLASSES[0]

    if is_junior_class(
        class_name
    ):
        department = "General"

    elif department not in DEPARTMENTS:
        department = "Science"

    conn = get_db()

    try:
        cur = conn.cursor()

        cur.execute(
            """
            SELECT *
            FROM subjects
            WHERE class_name = %s
              AND department = %s
              AND term = %s
            ORDER BY subject_name
            """,
            (
                class_name,
                department,
                term
            )
        )

        subject_rows = cur.fetchall()

    finally:
        conn.close()

    return render_template(
        "teacher.html",
        teacher_name=session.get(
            "teacher_name"
        ),
        subjects=subject_rows,
        class_name=class_name,
        department=department,
        term=term,
        classes=CLASSES,
        departments=DEPARTMENTS,
        terms=TERMS
    )


@app.route("/teacher/news")
@teacher_required
def teacher_news():

    conn = get_db()

    try:
        cur = conn.cursor()

        cur.execute(
            """
            SELECT *
            FROM news
            ORDER BY id DESC
            """
        )

        news_list = cur.fetchall()

    finally:
        conn.close()

    return render_template(
        "news.html",
        news=news_list,
        back_url=url_for("teacher"),
        back_text="Teacher Panel"
    )


@app.route("/admin")
@admin_required
def admin():

    section = request.args.get(
        "section",
        ""
    ).strip().lower()

    conn = get_db()

    try:
        cur = conn.cursor()

        cur.execute(
            "SELECT COUNT(*) AS count FROM students"
        )
        student_count = cur.fetchone()["count"]

        cur.execute(
            "SELECT COUNT(*) AS count FROM teachers"
        )
        teacher_count = cur.fetchone()["count"]

        cur.execute(
            "SELECT COUNT(*) AS count FROM subjects"
        )
        subject_count = cur.fetchone()["count"]

        students = []
        teachers = []
        subjects_list = []

        if section == "students":

            cur.execute(
                """
                SELECT
                    id,
                    name,
                    class_name,
                    department
                FROM students
                ORDER BY name
                """
            )

            students = cur.fetchall()

        elif section == "teachers":

            cur.execute(
                """
                SELECT
                    id,
                    name
                FROM teachers
                ORDER BY name
                """
            )

            teachers = cur.fetchall()

        elif section == "subjects":

            cur.execute(
                """
                SELECT *
                FROM subjects
                ORDER BY
                    class_name,
                    department,
                    term,
                    subject_name
                """
            )

            subjects_list = cur.fetchall()

    finally:
        conn.close()

    return render_template(
        "admin.html",
        section=section,
        students=students,
        teachers=teachers,
        subjects=subjects_list,
        student_count=student_count,
        teacher_count=teacher_count,
        subject_count=subject_count,
        classes=CLASSES,
        departments=DEPARTMENTS,
        terms=TERMS
    )


@app.route("/admin/dashboard")
@admin_required
def admin_dashboard():
    return redirect(
        url_for("admin")
    )


@app.route(
    "/admin/register-student",
    methods=["POST"]
)
@admin_required
def register_student():

    name = normalize_name(
        request.form.get(
            "name",
            ""
        )
    )

    class_name = request.form.get(
        "class_name",
        ""
    ).strip().upper()

    department = request.form.get(
        "department",
        ""
    ).strip()

    password = request.form.get(
        "password",
        ""
    )

    if not name or not class_name or not password:
        flash(
            "Please fill all required student fields."
        )

        return redirect(
            url_for("admin")
        )

    if class_name not in CLASSES:
        flash(
            "Invalid class selected."
        )

        return redirect(
            url_for("admin")
        )

    if is_junior_class(
        class_name
    ):

        department = "General"

    else:

        if department not in DEPARTMENTS:
            flash(
                "Senior students must have Science, Art or Commercial department."
            )

            return redirect(
                url_for("admin")
            )

    hashed_password = generate_password_hash(
        password
    )

    conn = get_db()

    try:

        cur = conn.cursor()

        cur.execute(
            """
            INSERT INTO students
                (
                    name,
                    class_name,
                    department,
                    password
                )
            VALUES
                (%s, %s, %s, %s)
            """,
            (
                name,
                class_name,
                department,
                hashed_password
            )
        )

        conn.commit()

        flash(
            f"Student {name} registered successfully."
        )

    except Exception as e:

        conn.rollback()

        print(
            "REGISTER STUDENT ERROR:",
            repr(e)
        )

        flash(
            "Could not register student. The name may already exist."
        )

    finally:
        conn.close()

    return redirect(
        url_for("admin")
    )


@app.route(
    "/admin/edit-student",
    methods=["POST"]
)
@admin_required
def edit_student():

    current_name = normalize_name(
        request.form.get(
            "current_name",
            ""
        )
    )

    new_name = normalize_name(
        request.form.get(
            "new_name",
            ""
        )
    )

    new_class = request.form.get(
        "class_name",
        ""
    ).strip().upper()

    new_department = request.form.get(
        "department",
        ""
    ).strip()

    new_password = request.form.get(
        "password",
        ""
    )

    if not current_name:
        flash(
            "Enter the student's currently registered name."
        )

        return redirect(
            url_for("admin")
        )

    conn = get_db()

    try:

        cur = conn.cursor()

        cur.execute(
            """
            SELECT *
            FROM students
            WHERE LOWER(TRIM(name))
                  = LOWER(TRIM(%s))
            LIMIT 1
            """,
            (current_name,)
        )

        student = cur.fetchone()

        if not student:
            flash(
                "Student not found."
            )

            return redirect(
                url_for("admin")
            )

        final_name = (
            new_name
            if new_name
            else student["name"]
        )

        final_class = (
            new_class
            if new_class
            else student["class_name"]
        )

        final_department = (
            new_department
            if new_department
            else student["department"]
        )

        if final_class not in CLASSES:
            flash(
                "Invalid class selected."
            )

            return redirect(
                url_for("admin")
            )

        if is_junior_class(
            final_class
        ):

            final_department = "General"

        else:

            if final_department not in DEPARTMENTS:

                flash(
                    "Senior students must have Science, Art or Commercial department."
                )

                return redirect(
                    url_for("admin")
                )

        if final_name.lower() != student["name"].lower():

            cur.execute(
                """
                SELECT id
                FROM students
                WHERE LOWER(TRIM(name))
                      = LOWER(TRIM(%s))
                  AND id <> %s
                LIMIT 1
                """,
                (
                    final_name,
                    student["id"]
                )
            )

            existing = cur.fetchone()

            if existing:
                flash(
                    "Another student already has that name."
                )

                return redirect(
                    url_for("admin")
                )

        if new_password:

            hashed_password = generate_password_hash(
                new_password
            )

            cur.execute(
                """
                UPDATE students
                SET
                    name = %s,
                    class_name = %s,
                    department = %s,
                    password = %s
                WHERE id = %s
                """,
                (
                    final_name,
                    final_class,
                    final_department,
                    hashed_password,
                    student["id"]
                )
            )

        else:

            cur.execute(
                """
                UPDATE students
                SET
                    name = %s,
                    class_name = %s,
                    department = %s
                WHERE id = %s
                """,
                (
                    final_name,
                    final_class,
                    final_department,
                    student["id"]
                )
            )

        conn.commit()

        flash(
            f"Student {final_name} updated successfully. Existing results were kept."
        )

    except Exception as e:

        conn.rollback()

        print(
            "EDIT STUDENT ERROR:",
            repr(e)
        )

        flash(
            "Could not update student."
        )

    finally:
        conn.close()

    return redirect(
        url_for("admin")
    )


@app.route(
    "/admin/delete-student",
    methods=["POST"]
)
@admin_required
def delete_student():

    student_name = normalize_name(
        request.form.get(
            "student_name",
            ""
        )
    )

    if not student_name:
        flash(
            "Enter the student's registered name."
        )

        return redirect(
            url_for("admin")
        )

    conn = get_db()

    try:

        cur = conn.cursor()

        cur.execute(
            """
            SELECT id, name
            FROM students
            WHERE LOWER(TRIM(name))
                  = LOWER(TRIM(%s))
            LIMIT 1
            """,
            (student_name,)
        )

        student = cur.fetchone()

        if not student:
            flash(
                "Student not found."
            )

            return redirect(
                url_for("admin")
            )

        cur.execute(
            """
            SELECT filename
            FROM results
            WHERE student_id = %s
            """,
            (student["id"],)
        )

        result_files = cur.fetchall()

        cur.execute(
            """
            DELETE FROM results
            WHERE student_id = %s
            """,
            (student["id"],)
        )

        cur.execute(
            """
            DELETE FROM students
            WHERE id = %s
            """,
            (student["id"],)
        )

        conn.commit()

        for result in result_files:

            delete_from_storage(
                storage_path(
                    result["filename"]
                )
            )

        flash(
            f"Student {student['name']} and their results were deleted."
        )

    except Exception as e:

        conn.rollback()

        print(
            "DELETE STUDENT ERROR:",
            repr(e)
        )

        flash(
            "Could not delete student."
        )

    finally:
        conn.close()

    return redirect(
        url_for("admin")
    )


@app.route(
    "/admin/register-teacher",
    methods=["POST"]
)
@admin_required
def register_teacher():

    name = normalize_name(
        request.form.get(
            "name",
            ""
        )
    )

    password = request.form.get(
        "password",
        ""
    )

    if not name or not password:

        flash(
            "Enter teacher name and password."
        )

        return redirect(
            url_for("admin")
        )

    hashed_password = generate_password_hash(
        password
    )

    conn = get_db()

    try:

        cur = conn.cursor()

        cur.execute(
            """
            INSERT INTO teachers
                (
                    name,
                    password
                )
            VALUES
                (%s, %s)
            """,
            (
                name,
                hashed_password
            )
        )

        conn.commit()

        flash(
            f"Teacher {name} registered successfully."
        )

    except Exception as e:

        conn.rollback()

        print(
            "REGISTER TEACHER ERROR:",
            repr(e)
        )

        flash(
            "Could not register teacher. The name may already exist."
        )

    finally:
        conn.close()

    return redirect(
        url_for("admin")
    )


@app.route(
    "/admin/delete-teacher",
    methods=["POST"]
)
@admin_required
def delete_teacher():

    teacher_name = normalize_name(
        request.form.get(
            "teacher_name",
            ""
        )
    )

    if not teacher_name:

        flash(
            "Enter the teacher's registered name."
        )

        return redirect(
            url_for("admin")
        )

    conn = get_db()

    try:

        cur = conn.cursor()

        cur.execute(
            """
            DELETE FROM teachers
            WHERE LOWER(TRIM(name))
                  = LOWER(TRIM(%s))
            """,
            (teacher_name,)
        )

        if cur.rowcount == 0:

            flash(
                "Teacher not found."
            )

        else:

            conn.commit()

            flash(
                f"Teacher {teacher_name} deleted."
            )

    except Exception as e:

        conn.rollback()

        print(
            "DELETE TEACHER ERROR:",
            repr(e)
        )

        flash(
            "Could not delete teacher."
        )

    finally:
        conn.close()

    return redirect(
        url_for("admin")
    )


@app.route(
    "/admin/add-subject",
    methods=["POST"]
)
@admin_required
def add_subject():

    subject_name = request.form.get(
        "subject_name",
        ""
    ).strip()

    subject_link = request.form.get(
        "subject_link",
        ""
    ).strip()

    class_name = request.form.get(
        "class_name",
        ""
    ).strip().upper()

    department = request.form.get(
        "department",
        ""
    ).strip()

    term = request.form.get(
        "term",
        ""
    ).strip()

    if (
        not subject_name
        or not subject_link
        or not class_name
        or not term
    ):

        flash(
            "Please fill all subject fields."
        )

        return redirect(
            url_for("admin")
        )

    if class_name not in CLASSES:

        flash(
            "Invalid class."
        )

        return redirect(
            url_for("admin")
        )

    if term not in TERMS:

        flash(
            "Invalid term."
        )

        return redirect(
            url_for("admin")
        )

    if is_junior_class(
        class_name
    ):

        department = "General"

    elif department not in DEPARTMENTS:

        flash(
            "Select a valid department."
        )

        return redirect(
            url_for("admin")
        )

    conn = get_db()

    try:

        cur = conn.cursor()

        cur.execute(
            """
            INSERT INTO subjects
                (
                    subject_name,
                    subject_link,
                    class_name,
                    department,
                    term
                )
            VALUES
                (%s, %s, %s, %s, %s)
            """,
            (
                subject_name,
                subject_link,
                class_name,
                department,
                term
            )
        )

        conn.commit()

        flash(
            "Subject/e-note added successfully."
        )

    except Exception as e:

        conn.rollback()

        print(
            "ADD SUBJECT ERROR:",
            repr(e)
        )

        flash(
            "Could not add subject."
        )

    finally:
        conn.close()

    return redirect(
        url_for("admin")
    )


@app.route(
    "/admin/delete-subject/<int:subject_id>",
    methods=["POST"]
)
@admin_required
def delete_subject(subject_id):

    conn = get_db()

    try:

        cur = conn.cursor()

        cur.execute(
            """
            DELETE FROM subjects
            WHERE id = %s
            """,
            (subject_id,)
        )

        conn.commit()

        flash(
            "Subject deleted successfully."
        )

    except Exception as e:

        conn.rollback()

        print(
            "DELETE SUBJECT ERROR:",
            repr(e)
        )

        flash(
            "Could not delete subject."
        )

    finally:
        conn.close()

    return redirect(
        url_for(
            "admin",
            section="subjects"
        )
    )


@app.route(
    "/admin/upload-result",
    methods=["POST"]
)
@admin_required
def upload_result():

    student_name = normalize_name(
        request.form.get(
            "student_name",
            ""
        )
    )

    term = request.form.get(
        "term",
        ""
    ).strip()

    uploaded_file = request.files.get(
        "result_file"
    )

    if not student_name:

        flash(
            "Enter the student's registered name."
        )

        return redirect(
            url_for("admin")
        )

    if term not in TERMS:

        flash(
            "Select a valid term."
        )

        return redirect(
            url_for("admin")
        )

    if (
        not uploaded_file
        or not uploaded_file.filename
    ):

        flash(
            "Please select a result file."
        )

        return redirect(
            url_for("admin")
        )

    original_filename = secure_filename(
        uploaded_file.filename
    )

    if not allowed_result_file(
        original_filename
    ):

        flash(
            "This result file type is not supported."
        )

        return redirect(
            url_for("admin")
        )

    file_bytes = uploaded_file.read()

    if not file_bytes:

        flash(
            "The uploaded file is empty."
        )

        return redirect(
            url_for("admin")
        )

    extension = os.path.splitext(
        original_filename
    )[1].lower()

    try:

        if extension == ".ods":
            validate_ods(
                file_bytes
            )

        elif extension in (
            ".xlsx",
            ".xlsm"
        ):
            load_workbook(
                io.BytesIO(file_bytes),
                read_only=True,
                data_only=True
            )

        elif extension == ".xls":

            try:
                import pandas as pd

                pd.read_excel(
                    io.BytesIO(file_bytes),
                    engine="xlrd"
                )

            except Exception as e:

                raise ValueError(
                    "The XLS file could not be validated."
                ) from e

    except Exception as e:

        print(
            "RESULT VALIDATION ERROR:",
            repr(e)
        )

        flash(
            "The uploaded result file is invalid or could not be read."
        )

        return redirect(
            url_for("admin")
        )

    conn = get_db()

    storage_file_path = None

    try:

        cur = conn.cursor()

        cur.execute(
            """
            SELECT id, name
            FROM students
            WHERE LOWER(TRIM(name))
                  = LOWER(TRIM(%s))
            LIMIT 1
            """,
            (student_name,)
        )

        student = cur.fetchone()

        if not student:

            flash(
                "Student not found."
            )

            return redirect(
                url_for("admin")
            )

        unique_filename = (
            f"{uuid.uuid4().hex}"
            f"_{original_filename}"
        )

        storage_file_path = storage_path(
            unique_filename
        )

        content_type = (
            guess_type(
                original_filename
            )[0]
            or "application/octet-stream"
        )

        upload_to_storage(
            storage_file_path,
            file_bytes,
            content_type
        )

        cur.execute(
            """
            INSERT INTO results
                (
                    student_id,
                    student_name,
                    filename,
                    original_filename,
                    term
                )
            VALUES
                (%s, %s, %s, %s, %s)
            """,
            (
                student["id"],
                student["name"],
                unique_filename,
                original_filename,
                term
            )
        )

        conn.commit()

        flash(
            f"Result uploaded successfully for {student['name']}."
        )

    except Exception as e:

        conn.rollback()

        print(
            "UPLOAD RESULT ERROR:",
            repr(e)
        )

        if storage_file_path:
            delete_from_storage(
                storage_file_path
            )

        flash(
            "The result could not be uploaded."
        )

    finally:
        conn.close()

    return redirect(
        url_for("admin")
    )


@app.route(
    "/admin/add-news",
    methods=["POST"]
)
@admin_required
def add_news():

    title = request.form.get(
        "title",
        ""
    ).strip()

    content = request.form.get(
        "content",
        ""
    ).strip()

    if not title or not content:

        flash(
            "Enter both news title and content."
        )

        return redirect(
            url_for("admin")
        )

    conn = get_db()

    try:

        cur = conn.cursor()

        cur.execute(
            """
            INSERT INTO news
                (
                    title,
                    content,
                    date,
                    admin
                )
            VALUES
                (%s, %s, %s, %s)
            """,
            (
                title,
                content,
                datetime.now().strftime(
                    "%Y-%m-%d %H:%M"
                ),
                session["admin"]
            )
        )

        conn.commit()

        flash(
            "School news published successfully."
        )

    except Exception as e:

        conn.rollback()

        print(
            "ADD NEWS ERROR:",
            repr(e)
        )

        flash(
            "Could not publish news."
        )

    finally:
        conn.close()

    return redirect(
        url_for("admin")
    )


@app.route(
    "/admin/delete-news/<int:news_id>",
    methods=["POST"]
)
@admin_required
def delete_news(news_id):

    conn = get_db()

    try:

        cur = conn.cursor()

        cur.execute(
            """
            DELETE FROM news
            WHERE id = %s
            """,
            (news_id,)
        )

        conn.commit()

        flash(
            "News deleted successfully."
        )

    except Exception as e:

        conn.rollback()

        print(
            "DELETE NEWS ERROR:",
            repr(e)
        )

        flash(
            "Could not delete news."
        )

    finally:
        conn.close()

    return redirect(
        url_for("admin")
    )


if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=int(
            os.environ.get(
                "PORT",
                5000
            )
        )
    )
